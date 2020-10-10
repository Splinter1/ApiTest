import openpyxl
import os

path = os.path.dirname(os.path.dirname(__file__))
filepath = path + "/data/"

# open_excel = openpyxl.load_workbook(filepath+"excel.xlsx")
# sheet_name = open_excel.sheetnames
# excel_value = open_excel[sheet_name[0]]
# print(excel_value)

# 获取一个单元格的数据
# print(excel_value.cell(1,2).value)
# print(excel_value.cell(2,3).value)
# print(excel_value.max_row)


class HandExcel:
    def load_excel(self):
        open_excel = openpyxl.load_workbook(filepath + "excel.xlsx")
        return open_excel

    def get_sheet_data(self,index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cell):
        data = self.get_sheet_data().cell(row, cell).value
        return data

    def get_rows(self):
        #   获取行数
        rows = self.get_sheet_data().max_row
        return rows

    def get_rows_value(self, row):
        # 获取一行的内容
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        # 写入数据
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(filepath + "excel.xlsx")

    def get_cols_value(self, col=None):
        # 获取一列内容
        cols_list = []
        if col is None:
            col = 'B'
        cols_list_data = self.get_sheet_data()[col]
        for i in cols_list_data:
            cols_list.append(i.value)
        return cols_list

    def get_rows_num(self, case_id):
        # 通过数据获取行号
        num = 1
        cols_data = self.get_cols_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1

    def get_excel_data(self):
        # 获取excel中所有数据  每行为list
        data_list = []
        rows = self.get_rows()
        for i in range(rows-1):
            data_list.append(self.get_rows_value(i+2))
        return data_list

    def get_cells_num(self, cell_name):
        # 根据字段名获取列号
        data_list = self.get_rows_value(1)
        cell_num = data_list.index(cell_name)
        return cell_num




if __name__ == '__main__':
    # print(HandExcel().get_cell_value(2, 12))
    # print(HandExcel().get_rows_value(2))
    # print(HandExcel().get_rows())
    # print(HandExcel().get_cols_value())
    # print(HandExcel().get_rows_num("addDevice3"))
    # print(HandExcel().get_excel_data())
    print(HandExcel().get_cells_num("是否执行"))